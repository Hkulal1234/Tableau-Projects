# -*- coding: utf-8 -*-
"""
Created on Thu Jan 30 05:16:25 2025

@author: harsh
"""
#data analysis on Amazon Sellter Data

import pandas as pd
import os
from openpyxl.styles import Font, PatternFill

# Load data from Excel
data = pd.read_excel(r'C:\Users\harsh\OneDrive\Desktop\Amazon-Seller-Data-Analysis\cleaned_orders.xlsx')

# Calculate total revenue
data['total_revenue'] = data['item_total'] + data['shipping_fee']

# Group Sales by Month
monthly_sales = data.groupby(data['order_date'].dt.to_period("M"))['total_revenue'].sum().reset_index()
monthly_sales['order_date'] = monthly_sales['order_date'].astype(str)  # Convert period to string

#calculate monthly sale growth

monthly_sales['MoM Growth (%)'] = monthly_sales['total_revenue'].pct_change().mul(100).round(2)

# Replace inf values with NaN after pct_change()
monthly_sales['MoM Growth (%)'] = monthly_sales['MoM Growth (%)'].replace([float('inf'), float('-inf')], float('nan'))
monthly_sales['Cumulative Sales'] = monthly_sales['total_revenue'].cumsum()

# Add the 'Total' row at the end of the dataframe
total_revenue = monthly_sales['total_revenue'].sum()
total_row = pd.DataFrame({'order_date': ['Total'], 'total_revenue': [total_revenue]})
total_cumulative_sales = monthly_sales['Cumulative Sales'].iloc[-1]  # Get the last cumulative sales value

# Calculate the average MoM Growth (ignoring NaN and infinity values)
average_mom_growth = monthly_sales['MoM Growth (%)'][1:].replace([float('inf'), float('-inf')], float('nan')).mean().round(2)

# Append 'Total' row once (with both total revenue and average MoM Growth)
total_row = pd.DataFrame({'order_date': ['Total'], 
                          'total_revenue': [total_revenue], 
                          'MoM Growth (%)': [average_mom_growth],
                          'Cumulative Sales': [total_cumulative_sales]})

# Append only once
monthly_sales = pd.concat([monthly_sales, total_row], ignore_index=True)


# Top 5 Selling Products
top_products = data.groupby('sku')['quantity'].sum().nlargest(5).reset_index()

# Orders by Payment Type
payment_type_count = data['cod'].value_counts().reset_index()
payment_type_count.columns = ['Payment Type', 'Order Count']

# Orders by Status
order_status_count = data['order_status'].value_counts().reset_index()
order_status_count.columns = ['Order Status', 'Count']

# Export Analysis Results to Excel
with pd.ExcelWriter(r'C:\Users\harsh\OneDrive\Desktop\Amazon-Seller-Data-Analysis\sales_analysis.xlsx', engine="openpyxl") as writer:
    # Write Monthly Sales to Excel
    monthly_sales.to_excel(writer, sheet_name="Monthly Sales", index=False)
    
    # Access the 'Monthly Sales' sheet for formatting
    workbook = writer.book
    worksheet = workbook["Monthly Sales"]
    
    # Get the last row index after writing all data to Excel (this should be the 'Total' row)
    last_row = worksheet.max_row  # Get the actual row count in Excel
    
    # Apply bold formatting to the 'Total' row
    worksheet.cell(row=worksheet.max_row, column=1).font = Font(bold=True)  # Bold 'Total'
    worksheet.cell(row=worksheet.max_row, column=2).font = Font(bold=True)  # Bold total revenue value
    worksheet.cell(row=worksheet.max_row, column=3).font = Font(bold=True)  # Bold MoM Growth
    worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)  # Bold Cumulative Sales
    
    # Get the last column index for 'MoM Growth (%)'
    mom_growth_col_idx = monthly_sales.columns.get_loc("MoM Growth (%)") + 1  # +1 since Excel is 1-based index

    # Apply red highlight for negative values in 'MoM Growth (%)'
    for row in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
        cell = worksheet.cell(row=row, column=mom_growth_col_idx)
        if isinstance(cell.value, (int, float)) and cell.value < 0:  # Check if value is negative
            cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Light Red Fill
    
    # Write Top Products to Excel
    top_products.to_excel(writer, sheet_name="Top Products", index=False)
    
    # Write Payment Type Analysis to Excel
    payment_type_count.to_excel(writer, sheet_name="Payment Type Analysis", index=False)
    
    # Write Order Status Breakdown to Excel
    order_status_count.to_excel(writer, sheet_name="Order Status Breakdown", index=False)

print("Analysis saved to sales_analysis.xlsx")

os.startfile(r'sales_analysis.xlsx') 